#!/usr/bin/env python3
"""Evaluate local 3D object-selection outputs against GT referents using an anchor table."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


class EvaluationError(Exception):
    """Raised when evaluation cannot proceed."""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Evaluate row_*.json local-3D predictions against GT referents with anchor/alias mapping."
    )
    parser.add_argument("--pred_dir", required=True, help="Directory containing row_*.json prediction files.")
    parser.add_argument("--gt_file", required=True, help="GT workbook or CSV file containing referents.")
    parser.add_argument("--anchor_csv", required=True, help="Anchor table TSV/CSV used for canonical label mapping.")
    parser.add_argument("--output_csv", required=True, help="Path to save per-row evaluation CSV.")
    parser.add_argument("--output_json", required=True, help="Path to save aggregate evaluation JSON.")
    parser.add_argument("--glob", default="row_*.json", help="Glob pattern for prediction JSON files.")
    parser.add_argument(
        "--summary_scope",
        choices=("all_gt", "predicted_rows"),
        default="all_gt",
        help="all_gt keeps the historical denominator; predicted_rows is useful for smoke tests. Default: all_gt.",
    )
    return parser.parse_args()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_label_key(value: Any) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"[\s_\-]+", "", text)
    return text


def split_alias_text(value: Any) -> List[str]:
    text = normalize_text(value)
    if not text:
        return []
    return [part.strip() for part in re.split(r"[,;/|]+", text) if part and part.strip()]


def detect_delimiter(path: Path) -> str:
    if path.suffix.lower() == ".tsv":
        return "\t"
    sample = path.read_text(encoding="utf-8", errors="ignore")[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t")
        return dialect.delimiter
    except csv.Error:
        return ","


def load_anchor_alias_map(path: Path) -> Tuple[Dict[str, str], List[str]]:
    if not path.exists() or not path.is_file():
        raise EvaluationError(f"Anchor table does not exist: {path}")

    delimiter = detect_delimiter(path)
    alias_map: Dict[str, str] = {}
    canonical_labels: List[str] = []
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        if reader.fieldnames is None:
            raise EvaluationError(f"Anchor table has no header: {path}")
        fieldnames = {name.lower(): name for name in reader.fieldnames}
        object_key = fieldnames.get("object_name") or fieldnames.get("物体名称")
        aliases_key = fieldnames.get("aliases") or fieldnames.get("alias") or fieldnames.get("object_aliases")
        if not object_key:
            raise EvaluationError(f"Anchor table is missing object_name column: {path}")

        for row in reader:
            canonical = normalize_text(row.get(object_key))
            if not canonical:
                continue
            canonical_labels.append(canonical)
            alias_map[normalize_label_key(canonical)] = canonical
            if aliases_key:
                for alias in split_alias_text(row.get(aliases_key)):
                    alias_map[normalize_label_key(alias)] = canonical
    if not canonical_labels:
        raise EvaluationError(f"No anchor rows found in: {path}")
    return alias_map, canonical_labels


def canonicalize_label(label: Any, alias_map: Dict[str, str]) -> str:
    raw_text = normalize_text(label)
    key = normalize_label_key(raw_text)
    if not key:
        return ""
    direct = alias_map.get(key, "")
    if direct:
        return direct

    fallback_patterns = (
        (r"^fork(\d+)$", "forklift{}"),
        (r"^direct(\d+)$", "truck{}"),
        (r"^cargo(\d+)$", "cargo{}"),
        (r"^building(\d+)$", "building{}"),
        (r"^helicopter(\d+)$", "helicopter{}"),
        (r"^person(\d+)$", "person{}"),
        (r"^chair(\d+)$", "chair{}"),
        (r"^desk(\d+)$", "desk{}"),
        (r"^laptop(\d+)$", "laptop{}"),
        (r"^pc(\d+)$", "pc{}"),
        (r"^lamp(\d+)$", "lamp{}"),
        (r"^file(\d+)$", "file{}"),
    )
    for pattern, template in fallback_patterns:
        match = re.match(pattern, key)
        if match:
            candidate = template.format(match.group(1))
            resolved = alias_map.get(normalize_label_key(candidate), "")
            if resolved:
                return resolved

    point_match = re.match(r"^point([abc])$", key)
    if point_match:
        candidate = f"point{point_match.group(1).upper()}"
        resolved = alias_map.get(normalize_label_key(candidate), "")
        if resolved:
            return resolved

    if key in {"platform1", "door1", "printer"}:
        resolved = alias_map.get(key, "")
        if resolved:
            return resolved

    return ""


def split_gt_referents(raw_value: Any) -> List[str]:
    text = normalize_text(raw_value)
    if not text:
        return []
    return [part.strip() for part in text.split(",") if part and part.strip()]


def find_column_index(headers: Sequence[str], candidates: Sequence[str], default_zero_based: int) -> int:
    candidate_set = {item.lower() for item in candidates}
    for index, value in enumerate(headers):
        if value in candidate_set:
            return index
    return default_zero_based


def load_gt_rows_xlsx(path: Path, alias_map: Dict[str, str]) -> List[Dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header_values = [normalize_text(cell).lower() for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]

    number_index = find_column_index(header_values, ("Number", "ID", "instruction_id"), 0)
    instruction_index = find_column_index(header_values, ("Instruction", "instruction_text"), 2)
    referent_index = find_column_index(header_values, ("Referents", "Referent", "target_description"), 3)

    rows: List[Dict[str, Any]] = []
    for excel_row_index, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if values is None:
            continue
        number = values[number_index] if len(values) > number_index else ""
        instruction = values[instruction_index] if len(values) > instruction_index else ""
        referents_raw = values[referent_index] if len(values) > referent_index else ""
        referent_list = split_gt_referents(referents_raw)
        canonical_referents = [canonicalize_label(item, alias_map) for item in referent_list]
        mapped_referents = [item for item in canonical_referents if item]
        unmapped_referents = [item for item, canon in zip(referent_list, canonical_referents) if not canon]
        rows.append(
            {
                "number": normalize_text(number) or str(len(rows) + 1),
                "excel_row_index": excel_row_index,
                "instruction": normalize_text(instruction),
                "referents_raw": normalize_text(referents_raw),
                "mapped_referents": mapped_referents,
                "unmapped_referents": unmapped_referents,
            }
        )
    return rows


def load_gt_rows_csv(path: Path, alias_map: Dict[str, str]) -> List[Dict[str, Any]]:
    delimiter = detect_delimiter(path)
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        if reader.fieldnames is None:
            raise EvaluationError(f"GT CSV has no header: {path}")
        fieldnames = {name.lower(): name for name in reader.fieldnames}
        number_key = fieldnames.get("number") or fieldnames.get("id") or fieldnames.get("instruction_id")
        instruction_key = fieldnames.get("instruction") or fieldnames.get("instruction_text")
        referent_key = fieldnames.get("referents") or fieldnames.get("referent") or fieldnames.get("target_description")
        if not referent_key:
            raise EvaluationError(f"Failed to find referent column in GT CSV: {path}")

        rows: List[Dict[str, Any]] = []
        for csv_row_index, row in enumerate(reader, start=2):
            referents_raw = row.get(referent_key, "")
            referent_list = split_gt_referents(referents_raw)
            canonical_referents = [canonicalize_label(item, alias_map) for item in referent_list]
            mapped_referents = [item for item in canonical_referents if item]
            unmapped_referents = [item for item, canon in zip(referent_list, canonical_referents) if not canon]
            rows.append(
                {
                    "number": normalize_text(row.get(number_key, "")) or str(len(rows) + 1),
                    "excel_row_index": csv_row_index,
                    "instruction": normalize_text(row.get(instruction_key, "")) if instruction_key else "",
                    "referents_raw": normalize_text(referents_raw),
                    "mapped_referents": mapped_referents,
                    "unmapped_referents": unmapped_referents,
                }
            )
    return rows


def load_gt_rows(path: Path, alias_map: Dict[str, str]) -> List[Dict[str, Any]]:
    if not path.exists() or not path.is_file():
        raise EvaluationError(f"GT file does not exist: {path}")
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        rows = load_gt_rows_xlsx(path, alias_map)
    elif suffix in {".csv", ".tsv"}:
        rows = load_gt_rows_csv(path, alias_map)
    else:
        raise EvaluationError(f"Unsupported GT file type: {path.suffix}")
    if not rows:
        raise EvaluationError(f"No GT rows were found in: {path}")
    return rows


def load_prediction_payload(path: Path) -> Dict[str, Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise EvaluationError(f"Failed to parse prediction JSON: {path} ({exc})") from exc


def unique_preserve_order(values: Iterable[str]) -> List[str]:
    result: List[str] = []
    seen = set()
    for value in values:
        text = normalize_text(value)
        if not text:
            continue
        key = normalize_label_key(text)
        if key in seen:
            continue
        seen.add(key)
        result.append(text)
    return result


def split_prediction_names(value: Any) -> List[str]:
    if isinstance(value, list):
        names: List[str] = []
        for item in value:
            if isinstance(item, dict):
                names.append(normalize_text(item.get("object_name") or item.get("selected_object_name")))
            else:
                names.append(normalize_text(item))
        return unique_preserve_order(names)
    if isinstance(value, str):
        return unique_preserve_order(part.strip() for part in re.split(r"[,;/|]+", value) if part.strip())
    return []


def find_prediction_labels(payload: Dict[str, Any]) -> List[str]:
    labels: List[str] = []
    for container_key in ("adjusted_response", "parsed_response"):
        container = payload.get(container_key)
        if isinstance(container, dict):
            labels.extend(split_prediction_names(container.get("selected_object_names")))
            labels.extend(split_prediction_names(container.get("selected_objects")))
            labels.append(normalize_text(container.get("selected_object_name")))

    resolved_rows = payload.get("resolved_object_rows")
    if isinstance(resolved_rows, list):
        labels.extend(
            normalize_text(item.get("object_name"))
            for item in resolved_rows
            if isinstance(item, dict)
        )

    resolved_row = payload.get("resolved_object_row")
    if isinstance(resolved_row, dict):
        labels.append(normalize_text(resolved_row.get("object_name")))

    return unique_preserve_order(labels)


def find_prediction_label(payload: Dict[str, Any]) -> str:
    labels = find_prediction_labels(payload)
    return labels[0] if labels else ""


def collect_prediction_files(pred_dir: Path, pattern: str) -> List[Path]:
    if not pred_dir.exists() or not pred_dir.is_dir():
        raise EvaluationError(f"Prediction directory does not exist: {pred_dir}")
    files = sorted(path for path in pred_dir.glob(pattern) if path.is_file())
    if not files:
        raise EvaluationError(f"No prediction files matched {pattern!r} in {pred_dir}")
    return files


def extract_row_index(path: Path, payload: Dict[str, Any]) -> int:
    row_index = payload.get("row_index")
    if isinstance(row_index, int):
        return row_index
    if isinstance(row_index, str) and row_index.strip().isdigit():
        return int(row_index.strip())
    match = re.search(r"row_(\d+)\.json$", path.name)
    if match:
        return int(match.group(1))
    raise EvaluationError(f"Could not determine row_index for prediction file: {path}")


def build_eval_rows(
    pred_files: Sequence[Path],
    gt_rows: Sequence[Dict[str, Any]],
    alias_map: Dict[str, str],
) -> List[Dict[str, Any]]:
    results: List[Dict[str, Any]] = []
    gt_count = len(gt_rows)

    for pred_path in pred_files:
        payload = load_prediction_payload(pred_path)
        row_index = extract_row_index(pred_path, payload)
        if row_index < 0 or row_index >= gt_count:
            raise EvaluationError(f"Prediction row_index={row_index} is out of range for GT rows ({gt_count}).")

        gt = gt_rows[row_index]
        predicted_raw_labels = find_prediction_labels(payload)
        predicted_canonical_labels = unique_preserve_order(
            canonicalize_label(label, alias_map) for label in predicted_raw_labels
        )
        mapped_gt = gt["mapped_referents"]
        predicted_set = set(predicted_canonical_labels)
        gt_set = set(mapped_gt)
        true_positive_set = predicted_set & gt_set
        false_positive_set = predicted_set - gt_set
        false_negative_set = gt_set - predicted_set
        hit = bool(true_positive_set)
        exact_match = bool(gt_set) and predicted_set == gt_set
        evaluable = len(mapped_gt) > 0
        precision = safe_ratio(len(true_positive_set), len(predicted_set))
        recall = safe_ratio(len(true_positive_set), len(gt_set))
        f1 = None
        if precision is not None and recall is not None and precision + recall > 0:
            f1 = 2 * precision * recall / (precision + recall)

        results.append(
            {
                "row_index": row_index,
                "instruction_number": gt["number"],
                "excel_row_index": gt["excel_row_index"],
                "event_id": normalize_text(payload.get("event_id")),
                "instruction": gt["instruction"],
                "gt_referents_raw": gt["referents_raw"],
                "gt_referents_mapped": ", ".join(mapped_gt),
                "gt_referents_unmapped": ", ".join(gt["unmapped_referents"]),
                "predicted_referent_raw": predicted_raw_labels[0] if predicted_raw_labels else "",
                "predicted_referent_mapped": predicted_canonical_labels[0] if predicted_canonical_labels else "",
                "predicted_referents_raw": ", ".join(predicted_raw_labels),
                "predicted_referents_mapped": ", ".join(predicted_canonical_labels),
                "true_positive_referents": ", ".join(sorted(true_positive_set)),
                "false_positive_referents": ", ".join(sorted(false_positive_set)),
                "false_negative_referents": ", ".join(sorted(false_negative_set)),
                "response_status": normalize_text(payload.get("response_status")),
                "used_storyboard_fallback": bool(payload.get("used_storyboard_fallback")),
                "match_success": hit,
                "exact_match": exact_match,
                "set_precision": precision,
                "set_recall": recall,
                "set_f1": f1,
                "evaluable": evaluable,
                "prediction_json": str(pred_path.resolve()),
            }
        )
    return sorted(results, key=lambda item: item["row_index"])


def write_eval_csv(path: Path, rows: Sequence[Dict[str, Any]]) -> None:
    fieldnames = [
        "row_index",
        "instruction_number",
        "excel_row_index",
        "event_id",
        "instruction",
        "gt_referents_raw",
        "gt_referents_mapped",
        "gt_referents_unmapped",
        "predicted_referent_raw",
        "predicted_referent_mapped",
        "predicted_referents_raw",
        "predicted_referents_mapped",
        "true_positive_referents",
        "false_positive_referents",
        "false_negative_referents",
        "response_status",
        "used_storyboard_fallback",
        "evaluable",
        "match_success",
        "exact_match",
        "set_precision",
        "set_recall",
        "set_f1",
        "prediction_json",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def safe_ratio(numerator: int, denominator: int) -> Optional[float]:
    return None if denominator <= 0 else numerator / denominator


def summarize(rows: Sequence[Dict[str, Any]], gt_rows: Sequence[Dict[str, Any]], summary_scope: str = "all_gt") -> Dict[str, Any]:
    matched_rows = [row for row in rows if row["match_success"]]
    exact_match_rows = [row for row in rows if row.get("exact_match")]
    evaluable_rows = [row for row in rows if row["evaluable"]]
    matched_evaluable_rows = [row for row in evaluable_rows if row["match_success"]]
    exact_match_evaluable_rows = [row for row in evaluable_rows if row.get("exact_match")]
    fallback_rows = [row for row in rows if row["used_storyboard_fallback"]]
    ok_rows = [row for row in rows if row["response_status"] == "ok"]

    selected_gt_rows = list(gt_rows)
    if summary_scope == "predicted_rows":
        selected_indices = {int(row["row_index"]) for row in rows}
        selected_gt_rows = [gt_rows[index] for index in sorted(selected_indices) if 0 <= index < len(gt_rows)]

    unmapped_counter = Counter()
    unsupported_gt_row_count = 0
    for gt in selected_gt_rows:
        if gt["unmapped_referents"]:
            unsupported_gt_row_count += 1
        for label in gt["unmapped_referents"]:
            unmapped_counter[normalize_text(label)] += 1

    predicted_counter = Counter()
    matched_counter = Counter()
    tp_total = 0
    fp_total = 0
    fn_total = 0
    per_row_precision_values: List[float] = []
    per_row_recall_values: List[float] = []
    per_row_f1_values: List[float] = []

    for row in rows:
        predicted_labels = split_gt_referents(row.get("predicted_referents_mapped"))
        tp_labels = split_gt_referents(row.get("true_positive_referents"))
        fp_labels = split_gt_referents(row.get("false_positive_referents"))
        fn_labels = split_gt_referents(row.get("false_negative_referents"))
        predicted_counter.update(predicted_labels)
        matched_counter.update(tp_labels)
        tp_total += len(tp_labels)
        fp_total += len(fp_labels)
        fn_total += len(fn_labels)
        if row["evaluable"]:
            if row.get("set_precision") is not None:
                per_row_precision_values.append(float(row["set_precision"]))
            if row.get("set_recall") is not None:
                per_row_recall_values.append(float(row["set_recall"]))
            if row.get("set_f1") is not None:
                per_row_f1_values.append(float(row["set_f1"]))

    total_gt_rows = len(selected_gt_rows)
    evaluable_row_count = len(evaluable_rows)
    match_count_overall = len(matched_rows)
    match_count_evaluable_only = len(matched_evaluable_rows)
    micro_precision = safe_ratio(tp_total, tp_total + fp_total)
    micro_recall = safe_ratio(tp_total, tp_total + fn_total)
    micro_f1 = None
    if micro_precision is not None and micro_recall is not None and micro_precision + micro_recall > 0:
        micro_f1 = 2 * micro_precision * micro_recall / (micro_precision + micro_recall)

    def average(values: Sequence[float]) -> Optional[float]:
        return None if not values else sum(values) / len(values)

    return {
        "total_gt_rows": total_gt_rows,
        "total_prediction_files": len(rows),
        "response_status_ok_count": len(ok_rows),
        "storyboard_fallback_count": len(fallback_rows),
        "match_count_overall": match_count_overall,
        "overall_accuracy": safe_ratio(match_count_overall, total_gt_rows),
        "exact_match_count_overall": len(exact_match_rows),
        "exact_match_accuracy_overall": safe_ratio(len(exact_match_rows), total_gt_rows),
        "evaluable_row_count": evaluable_row_count,
        "evaluable_coverage": safe_ratio(evaluable_row_count, total_gt_rows),
        "unevaluable_row_count": total_gt_rows - evaluable_row_count,
        "match_count_evaluable_only": match_count_evaluable_only,
        "mapped_only_accuracy": safe_ratio(match_count_evaluable_only, evaluable_row_count),
        "exact_match_count_evaluable_only": len(exact_match_evaluable_rows),
        "exact_match_accuracy_evaluable_only": safe_ratio(len(exact_match_evaluable_rows), evaluable_row_count),
        "set_true_positive_count": tp_total,
        "set_false_positive_count": fp_total,
        "set_false_negative_count": fn_total,
        "micro_precision": micro_precision,
        "micro_recall": micro_recall,
        "micro_f1": micro_f1,
        "macro_precision": average(per_row_precision_values),
        "macro_recall": average(per_row_recall_values),
        "macro_f1": average(per_row_f1_values),
        "unsupported_gt_row_count": unsupported_gt_row_count,
        "unsupported_gt_referent_mention_count": sum(unmapped_counter.values()),
        "unsupported_gt_distinct_label_count": len(unmapped_counter),
        "unsupported_gt_referent_counts": dict(sorted(unmapped_counter.items(), key=lambda item: (-item[1], item[0]))),
        "predicted_referent_counts": dict(sorted(predicted_counter.items(), key=lambda item: (-item[1], item[0]))),
        "matched_referent_counts": dict(sorted(matched_counter.items(), key=lambda item: (-item[1], item[0]))),
    }


def write_summary_json(path: Path, summary: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    args = parse_args()

    pred_dir = Path(args.pred_dir).resolve()
    gt_file = Path(args.gt_file).resolve()
    anchor_csv = Path(args.anchor_csv).resolve()
    output_csv = Path(args.output_csv).resolve()
    output_json = Path(args.output_json).resolve()

    alias_map, _ = load_anchor_alias_map(anchor_csv)
    gt_rows = load_gt_rows(gt_file, alias_map)
    pred_files = collect_prediction_files(pred_dir, args.glob)
    eval_rows = build_eval_rows(pred_files, gt_rows, alias_map)
    summary = summarize(eval_rows, gt_rows, args.summary_scope)
    summary["summary_scope"] = args.summary_scope

    write_eval_csv(output_csv, eval_rows)
    write_summary_json(output_json, summary)

    print(f"Saved per-row evaluation CSV to: {output_csv}")
    print(f"Saved aggregate evaluation JSON to: {output_json}")
    print(
        f"Overall accuracy: {summary['match_count_overall']}/{summary['total_gt_rows']} = "
        f"{summary['overall_accuracy']}"
    )
    print(
        f"Mapped-only accuracy: {summary['match_count_evaluable_only']}/{summary['evaluable_row_count']} = "
        f"{summary['mapped_only_accuracy']}"
    )
    print(
        f"Set micro F1: precision={summary['micro_precision']} "
        f"recall={summary['micro_recall']} f1={summary['micro_f1']}"
    )
    print(
        f"Exact-set accuracy: {summary['exact_match_count_evaluable_only']}/{summary['evaluable_row_count']} = "
        f"{summary['exact_match_accuracy_evaluable_only']}"
    )
    unsupported = summary["unsupported_gt_referent_counts"]
    if unsupported:
        print("Unsupported GT referents detected (not present in current anchor-label space):")
        for label, count in unsupported.items():
            print(f"  - {label}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
