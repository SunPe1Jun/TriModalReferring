#!/usr/bin/env python3
"""Evaluate scene object-selection JSON outputs against referents in an xlsx workbook."""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook


class EvaluationError(Exception):
    """Raised when evaluation cannot proceed."""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Evaluate scene 3D referent predictions against xlsx referents.")
    parser.add_argument("--pred_dir", required=True, help="Directory containing row_*.json prediction files.")
    parser.add_argument("--gt_xlsx", required=True, help="Path to scene1.xlsx.")
    parser.add_argument("--output_csv", required=True, help="Path to save per-row evaluation CSV.")
    parser.add_argument("--output_json", required=True, help="Path to save aggregate evaluation JSON.")
    parser.add_argument(
        "--glob",
        default="row_*.json",
        help="Glob pattern for prediction JSON files inside pred_dir. Default: row_*.json",
    )
    return parser.parse_args()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def canonicalize_anchor_label(label: str) -> str:
    text = normalize_text(label).strip().lower()
    text = re.sub(r"[\s\-]+", "", text)
    if not text:
        return ""

    direct_match_patterns = (
        r"^(cargo\d+)$",
        r"^(building\d+)$",
        r"^(helicopter\d+)$",
        r"^(person\d+)$",
        r"^(truck\d+)$",
        r"^(forklift\d+)$",
        r"^(platform1)$",
        r"^(chair\d+)$",
        r"^(desk\d+)$",
        r"^(laptop\d+)$",
        r"^(pc\d+)$",
        r"^(lamp\d+)$",
        r"^(file\d+)$",
        r"^(point[a-c])$",
        r"^(door1)$",
        r"^(printer)$",
    )
    for pattern in direct_match_patterns:
        match = re.match(pattern, text)
        if match:
            return match.group(1)

    short_map_patterns = (
        (r"^fork(\d+)$", "forklift{}"),
        (r"^direct(\d+)$", "truck{}"),
        (r"^cargo(\d+)$", "cargo{}"),
        (r"^building(\d+)$", "building{}"),
        (r"^helicopter(\d+)$", "helicopter{}"),
        (r"^person(\d+)$", "person{}"),
        (r"^chair(\d+)$", "chair{}"),
        (r"^desk(\d+)$", "desk{}"),
        (r"^laptop(\d+)$", "laptop{}"),
        (r"^pc(\d+)$", "pc{}"),
        (r"^lamp(\d+)$", "lamp{}"),
        (r"^file(\d+)$", "file{}"),
    )
    for pattern, template in short_map_patterns:
        match = re.match(pattern, text)
        if match:
            return template.format(match.group(1))
    point_match = re.match(r"^point([abc])$", text)
    if point_match:
        return f"point{point_match.group(1)}"
    if text == "door1":
        return "door1"
    if text == "printer":
        return "printer"

    return ""


def split_gt_referents(raw_value: Any) -> List[str]:
    text = normalize_text(raw_value)
    if not text:
        return []
    return [part.strip() for part in text.split(",") if part and part.strip()]


def load_gt_rows(xlsx_path: Path) -> List[Dict[str, Any]]:
    if not xlsx_path.exists() or not xlsx_path.is_file():
        raise EvaluationError(f"GT xlsx does not exist: {xlsx_path}")

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    header_values = []
    for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)):
        header_values.append(normalize_text(cell).lower())

    def find_column_index(candidates: Sequence[str], default_zero_based: int) -> int:
        candidate_set = {item.lower() for item in candidates}
        for index, value in enumerate(header_values):
            if value in candidate_set:
                return index
        return default_zero_based

    number_index = find_column_index(("Number", "ID"), 0)
    instruction_index = find_column_index(("Instruction",), 2)
    referent_index = find_column_index(("Referents", "Referent"), 3)

    rows: List[Dict[str, Any]] = []
    for excel_row_index, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if values is None:
            continue
        number = values[number_index] if len(values) > number_index else ""
        instruction = values[instruction_index] if len(values) > instruction_index else ""
        referents_raw = values[referent_index] if len(values) > referent_index else ""
        referent_list = split_gt_referents(referents_raw)
        canonical_referents = [canonicalize_anchor_label(item) for item in referent_list]
        mapped_referents = [item for item in canonical_referents if item]
        unmapped_referents = [item for item, canon in zip(referent_list, canonical_referents) if not canon]
        rows.append(
            {
                "number": normalize_text(number) or str(len(rows) + 1),
                "excel_row_index": excel_row_index,
                "instruction": normalize_text(instruction),
                "referents_raw": normalize_text(referents_raw),
                "referents_list": referent_list,
                "canonical_referents": canonical_referents,
                "mapped_referents": mapped_referents,
                "unmapped_referents": unmapped_referents,
            }
        )
    if not rows:
        raise EvaluationError(f"No GT rows were found in: {xlsx_path}")
    return rows


def load_prediction_payload(path: Path) -> Dict[str, Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise EvaluationError(f"Failed to parse prediction JSON: {path} ({exc})") from exc


def find_prediction_label(payload: Dict[str, Any]) -> str:
    for container_key in ("adjusted_response", "parsed_response", "resolved_object_row"):
        container = payload.get(container_key)
        if isinstance(container, dict):
            if container_key == "resolved_object_row":
                label = normalize_text(container.get("object_name"))
            else:
                label = normalize_text(container.get("selected_object_name"))
            if label:
                return label
    return ""


def collect_prediction_files(pred_dir: Path, pattern: str) -> List[Path]:
    if not pred_dir.exists() or not pred_dir.is_dir():
        raise EvaluationError(f"Prediction directory does not exist: {pred_dir}")
    files = sorted(path for path in pred_dir.glob(pattern) if path.is_file())
    if not files:
        raise EvaluationError(f"No prediction files matched {pattern!r} in {pred_dir}")
    return files


def extract_row_index(path: Path, payload: Dict[str, Any]) -> int:
    row_index = payload.get("row_index")
    if isinstance(row_index, int):
        return row_index
    if isinstance(row_index, str) and row_index.strip().isdigit():
        return int(row_index.strip())
    match = re.search(r"row_(\d+)\.json$", path.name)
    if match:
        return int(match.group(1))
    raise EvaluationError(f"Could not determine row_index for prediction file: {path}")


def build_eval_rows(pred_files: Sequence[Path], gt_rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    results: List[Dict[str, Any]] = []
    gt_count = len(gt_rows)

    for pred_path in pred_files:
        payload = load_prediction_payload(pred_path)
        row_index = extract_row_index(pred_path, payload)
        if row_index < 0 or row_index >= gt_count:
            raise EvaluationError(f"Prediction row_index={row_index} is out of range for GT rows ({gt_count}).")

        gt = gt_rows[row_index]
        predicted_raw = find_prediction_label(payload)
        predicted_canonical = canonicalize_anchor_label(predicted_raw)
        mapped_gt = gt["mapped_referents"]
        hit = bool(predicted_canonical and predicted_canonical in mapped_gt)
        evaluable = len(mapped_gt) > 0

        results.append(
            {
                "row_index": row_index,
                "instruction_number": gt["number"],
                "excel_row_index": gt["excel_row_index"],
                "event_id": normalize_text(payload.get("event_id")),
                "instruction": gt["instruction"],
                "gt_referents_raw": gt["referents_raw"],
                "gt_referents_mapped": ", ".join(mapped_gt),
                "gt_referents_unmapped": ", ".join(gt["unmapped_referents"]),
                "predicted_referent_raw": predicted_raw,
                "predicted_referent_mapped": predicted_canonical,
                "response_status": normalize_text(payload.get("response_status")),
                "used_storyboard_fallback": bool(payload.get("used_storyboard_fallback")),
                "match_success": hit,
                "evaluable": evaluable,
                "prediction_json": str(pred_path.resolve()),
            }
        )
    return sorted(results, key=lambda item: item["row_index"])


def write_eval_csv(path: Path, rows: Sequence[Dict[str, Any]]) -> None:
    fieldnames = [
        "row_index",
        "instruction_number",
        "excel_row_index",
        "event_id",
        "instruction",
        "gt_referents_raw",
        "gt_referents_mapped",
        "gt_referents_unmapped",
        "predicted_referent_raw",
        "predicted_referent_mapped",
        "response_status",
        "used_storyboard_fallback",
        "evaluable",
        "match_success",
        "prediction_json",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def safe_ratio(numerator: int, denominator: int) -> Optional[float]:
    if denominator <= 0:
        return None
    return numerator / denominator


def summarize(rows: Sequence[Dict[str, Any]], gt_rows: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
    total_predictions = len(rows)
    total_gt_rows = len(gt_rows)
    matched_rows = [row for row in rows if row["match_success"]]
    evaluable_rows = [row for row in rows if row["evaluable"]]
    matched_evaluable_rows = [row for row in evaluable_rows if row["match_success"]]
    fallback_rows = [row for row in rows if row["used_storyboard_fallback"]]
    ok_rows = [row for row in rows if row["response_status"] == "ok"]

    unmapped_counter = Counter()
    for gt in gt_rows:
        for label in gt["unmapped_referents"]:
            unmapped_counter[normalize_text(label)] += 1

    predicted_counter = Counter(row["predicted_referent_mapped"] for row in rows if row["predicted_referent_mapped"])
    matched_counter = Counter(row["predicted_referent_mapped"] for row in matched_rows if row["predicted_referent_mapped"])

    missing_prediction_rows = sorted(set(range(total_gt_rows)) - {row["row_index"] for row in rows})

    return {
        "total_gt_rows": total_gt_rows,
        "total_prediction_files": total_predictions,
        "missing_prediction_row_indices": missing_prediction_rows,
        "response_status_ok_count": len(ok_rows),
        "storyboard_fallback_count": len(fallback_rows),
        "match_count_overall": len(matched_rows),
        "overall_accuracy": safe_ratio(len(matched_rows), total_gt_rows),
        "evaluable_row_count": len(evaluable_rows),
        "match_count_evaluable_only": len(matched_evaluable_rows),
        "mapped_only_accuracy": safe_ratio(len(matched_evaluable_rows), len(evaluable_rows)),
        "unsupported_gt_referent_counts": dict(sorted(unmapped_counter.items(), key=lambda item: (-item[1], item[0]))),
        "predicted_referent_counts": dict(sorted(predicted_counter.items(), key=lambda item: (-item[1], item[0]))),
        "matched_referent_counts": dict(sorted(matched_counter.items(), key=lambda item: (-item[1], item[0]))),
    }


def main() -> int:
    args = parse_args()
    try:
        pred_dir = Path(args.pred_dir).expanduser().resolve()
        gt_xlsx = Path(args.gt_xlsx).expanduser().resolve()
        pred_files = collect_prediction_files(pred_dir, args.glob)
        gt_rows = load_gt_rows(gt_xlsx)
        eval_rows = build_eval_rows(pred_files, gt_rows)

        output_csv = Path(args.output_csv).expanduser().resolve()
        output_json = Path(args.output_json).expanduser().resolve()
        write_eval_csv(output_csv, eval_rows)

        summary = summarize(eval_rows, gt_rows)
        output_json.parent.mkdir(parents=True, exist_ok=True)
        output_json.write_text(
            json.dumps(
                {
                    "summary": summary,
                    "notes": {
                        "match_rule": "Prediction counts as success if predicted selected_object_name matches any canonicalized GT referent for that row.",
                        "mapping_examples": {
                            "Fork2": "forklift2",
                            "Direct2": "truck2",
                            "Cargo1": "cargo1",
                        },
                        "mapped_only_accuracy_definition": "Accuracy computed only on rows where at least one GT referent can be mapped into the current anchor-label space.",
                    },
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        print(f"Saved per-row evaluation CSV to: {output_csv}")
        print(f"Saved aggregate evaluation JSON to: {output_json}")
        print(f"Overall accuracy: {summary['match_count_overall']}/{summary['total_gt_rows']} = {summary['overall_accuracy']}")
        print(
            f"Mapped-only accuracy: {summary['match_count_evaluable_only']}/{summary['evaluable_row_count']} = {summary['mapped_only_accuracy']}"
        )
        unsupported = summary["unsupported_gt_referent_counts"]
        if unsupported:
            print("Unsupported GT referents detected (not present in current anchor-label space):")
            for label, count in list(unsupported.items())[:20]:
                print(f"  - {label}: {count}")
        return 0
    except EvaluationError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
