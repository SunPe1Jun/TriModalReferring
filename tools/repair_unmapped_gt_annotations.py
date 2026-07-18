#!/usr/bin/env python3
"""Apply the human-confirmed GT and anchor repairs for the 4,000-row corpus."""

from __future__ import annotations

import argparse
import copy
import csv
import json
import math
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Sequence, Tuple

import openpyxl


SCENE2_FINAL = {
    260: ("Laptop1", "laptop1"),
    265: ("Printer", "printer"),
    335: ("Printer", "printer"),
    361: ("Laptop1, Printer", "laptop1, printer"),
    374: ("Laptop1", "laptop1"),
    498: ("desk4, file4, desk1", "desk4, file4, desk1"),
    751: ("Printer, Drawer2", "printer, drawer2"),
    792: ("Printer, Drawer2", "printer, drawer2"),
}

REFERENT_REPAIRS = {
    "scene4_room1": {138: "rug, red sofa"},
    "scene4_room3": {19: "room1"},
    "scene5": {
        9: "drone 5, seawall",
        35: "drone 5, coastal road",
        58: "drone 4, coastal road",
        131: "drone 5, coastal road",
    },
}

CONFIRMED_CANONICAL = {
    "scene2": {
        260: ["laptop1"], 265: ["printer"], 335: ["printer"],
        361: ["laptop1", "printer"], 374: ["laptop1"],
        498: ["desk4", "file4", "desk1"],
        751: ["printer", "drawer2"], 792: ["printer", "drawer2"],
    },
    "scene4_room1": {
        15: ["lamp"], 20: ["cushion"], 33: ["stove"], 34: ["armchair"],
        37: ["cushion"], 44: ["painting"], 45: ["ceiling_lamp"], 64: ["lamp"],
        70: ["lamp", "sofa"], 76: ["cushion", "armchair"], 94: ["armchair"],
        99: ["stove", "sofa"], 130: ["lamp", "ceiling_lamp"], 138: ["rug", "sofa"],
    },
    "scene4_room3": {19: ["room1"]},
    "scene4_room4": {61: ["vase"]},
    "scene5": {
        9: ["drone5", "seawall"], 35: ["drone5", "coastal_road"],
        58: ["drone4", "coastal_road"], 131: ["drone5", "coastal_road"],
    },
}

INSTRUCTION_CHECKS = {
    ("scene2", 260): "Confirm that laptop",
    ("scene2", 265): "Confirm that printer",
    ("scene2", 335): "Review that printer",
    ("scene2", 361): "Slide that laptop",
    ("scene2", 374): "Inspect that laptop",
    ("scene2", 498): "Confirm that desk",
    ("scene2", 569): "Close that drawer",
    ("scene2", 751): "Confirm that printer",
    ("scene2", 792): "Inspect that printer",
    ("scene4_room1", 138): "Fold the rug",
    ("scene4_room3", 19): "area in front of room1",
    ("scene4_room4", 61): "dining table centerpiece",
    ("scene5", 9): "seawall area",
    ("scene5", 35): "coastal road",
    ("scene5", 58): "road curve",
    ("scene5", 131): "coastal road",
}

ALIASES = {
    "scene2": {
        "laptop1": ["Laptop3", "laptop3", "laptop 3", "laptop 1"],
        "printer": ["Printer1", "printer1", "printer 1"],
        "drawer1": ["Drawer1", "drawer 1"],
        "drawer2": ["Drawer2", "drawer 2"],
    },
    "scene4_room1": {
        "lamp": ["floor lamp"],
        "cushion": ["sofa cushion", "cushions"],
        "armchair": ["blue armchair"],
        "painting": ["wall picture"],
        "sofa": ["red sofa"],
        "stove": ["heater"],
        "ceiling_lamp": ["ceiling", "ceiling light", "ceiling lamp"],
    },
    "scene4_room3": {
        "room1": ["room1_front_area", "area in front of room1", "front of room1"],
    },
    "scene4_room4": {
        "vase": ["centerpiece", "dining table centerpiece"],
    },
}


def normalize(value: Any) -> str:
    return "" if value is None else str(value).strip()


def label_key(value: Any) -> str:
    return re.sub(r"[\s_-]+", "", normalize(value).lower())


def headers(ws: Any) -> Dict[str, int]:
    return {normalize(ws.cell(1, col).value): col for col in range(1, ws.max_column + 1) if normalize(ws.cell(1, col).value)}


def verify_instruction(scene: str, row_index: int, value: Any) -> None:
    expected = INSTRUCTION_CHECKS.get((scene, row_index))
    if expected and expected.lower() not in normalize(value).lower():
        raise RuntimeError(f"instruction mismatch for {scene}::{row_index}: expected {expected!r}, got {value!r}")


def record_change(changes: List[Dict[str, Any]], path: Path, sheet: str, row: int, column: str, old: Any, new: Any) -> None:
    if old == new:
        return
    changes.append({
        "file": str(path), "sheet": sheet, "excel_row": row, "column": column,
        "old_value": old, "new_value": new,
    })


def set_cell(ws: Any, row: int, col: int, value: Any, changes: List[Dict[str, Any]], path: Path) -> None:
    cell = ws.cell(row, col)
    record_change(changes, path, ws.title, row, normalize(ws.cell(1, col).value), cell.value, value)
    cell.value = value


def copy_cell_style(source: Any, target: Any) -> None:
    target._style = copy.copy(source._style)
    target.number_format = source.number_format
    target.font = copy.copy(source.font)
    target.fill = copy.copy(source.fill)
    target.border = copy.copy(source.border)
    target.alignment = copy.copy(source.alignment)
    target.protection = copy.copy(source.protection)


def repair_workbook(repo_root: Path, scene: str, filename: str, changes: List[Dict[str, Any]]) -> Path:
    path = repo_root / "data" / filename
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[wb.sheetnames[0]]
    cols = headers(ws)
    instruction_col = cols["Instruction"]
    referent_col = cols["Referents"]

    if scene == "scene2":
        for row_index, (display, canonical) in SCENE2_FINAL.items():
            excel_row = row_index + 2
            verify_instruction(scene, row_index, ws.cell(excel_row, instruction_col).value)
            set_cell(ws, excel_row, referent_col, display, changes, path)
            for column_name, value in (
                ("Referents_Cleaned", display),
                ("Referents_AnchorMapped", canonical),
                ("Referents_Changed", "true"),
                ("Referents_Unparsed", None),
            ):
                if column_name in cols:
                    set_cell(ws, excel_row, cols[column_name], value, changes, path)
        verify_instruction(scene, 569, ws.cell(571, instruction_col).value)

        object_col = cols["object_name"]
        existing = {label_key(ws.cell(row, object_col).value): row for row in range(2, ws.max_row + 1) if normalize(ws.cell(row, object_col).value)}
        if "drawer2" in existing:
            anchor_row = existing["drawer2"]
        else:
            anchor_row = next(
                row for row in range(2, ws.max_row + 1)
                if all(ws.cell(row, col).value in (None, "") for col in range(object_col, object_col + 4))
            )
            style_row = max(row for row in existing.values() if row < anchor_row)
            for offset in range(4):
                copy_cell_style(ws.cell(style_row, object_col + offset), ws.cell(anchor_row, object_col + offset))
        for offset, value in enumerate(("drawer2", 0.638, -1.227, 5.241)):
            set_cell(ws, anchor_row, object_col + offset, value, changes, path)
    else:
        for row_index, value in REFERENT_REPAIRS.get(scene, {}).items():
            excel_row = row_index + 2
            verify_instruction(scene, row_index, ws.cell(excel_row, instruction_col).value)
            set_cell(ws, excel_row, referent_col, value, changes, path)
        for row_index in CONFIRMED_CANONICAL.get(scene, {}):
            verify_instruction(scene, row_index, ws.cell(row_index + 2, instruction_col).value)

    if scene == "scene4_room3":
        # The corrupt annotation also spilled into the side-by-side anchor area.
        excel_row = 21
        object_col = cols["object_name"]
        for col in range(object_col, min(object_col + 4, ws.max_column + 1)):
            set_cell(ws, excel_row, col, None, changes, path)

    wb.save(path)
    return path


def parse_anchor_table(path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    lines = path.read_text(encoding="utf-8-sig").splitlines()
    header = lines[0].split("\t")
    rows: List[Dict[str, str]] = []
    for line in lines[1:]:
        if not line.strip():
            continue
        parts = line.split("\t")
        if len(parts) < len(header) and header[-1] == "aliases":
            parts.extend([""] * (len(header) - len(parts)))
        if len(parts) == len(header):
            rows.append(dict(zip(header, parts)))
            continue
        if parts[0].strip() == "drawer1":
            values = re.findall(r"[-+]?\d+(?:\.\d+)?", " ".join(parts[1:]))
            if len(values) != 3:
                raise RuntimeError(f"cannot repair drawer1 row in {path}: {line!r}")
            rows.append(dict(zip(header[:4], ["drawer1", *values])))
            continue
        raise RuntimeError(f"malformed anchor row in {path}: {line!r}")
    return header, rows


def merge_aliases(existing: Any, additions: Iterable[str]) -> str:
    values = [part.strip() for part in re.split(r"[,;/|]+", normalize(existing)) if part.strip()]
    seen = {label_key(value) for value in values}
    for value in additions:
        if label_key(value) not in seen:
            values.append(value)
            seen.add(label_key(value))
    return ", ".join(values)


def repair_anchor_table(repo_root: Path, scene: str, changes: List[Dict[str, Any]]) -> Path:
    path = repo_root / "data" / f"{scene}_anchor_table.tsv"
    header, rows = parse_anchor_table(path)
    if "aliases" not in header:
        header.append("aliases")
    by_key = {label_key(row.get("object_name")): row for row in rows}
    if scene == "scene2":
        drawer1 = by_key.get("drawer1")
        if drawer1 is None:
            raise RuntimeError("drawer1 missing after parsing scene2 anchor table")
        drawer1.update({"x_world": "0.097", "y_world": "-1.2", "z_world": "2.694"})
        if "drawer2" not in by_key:
            row = {name: "" for name in header}
            row.update({"object_name": "drawer2", "x_world": "0.638", "y_world": "-1.227", "z_world": "5.241"})
            rows.append(row)
            by_key["drawer2"] = row
    for canonical, aliases in ALIASES.get(scene, {}).items():
        row = by_key.get(label_key(canonical))
        if row is None:
            raise RuntimeError(f"anchor {canonical!r} missing in {path}")
        old = normalize(row.get("aliases"))
        new = merge_aliases(old, aliases)
        if old != new:
            changes.append({"file": str(path), "sheet": "", "excel_row": "", "column": f"aliases:{canonical}", "old_value": old, "new_value": new})
        row["aliases"] = new

    for row in rows:
        for field in ("x_world", "y_world", "z_world"):
            value = float(normalize(row.get(field)))
            if not math.isfinite(value):
                raise RuntimeError(f"nonfinite {field} for {row.get('object_name')} in {path}")
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=header, delimiter="\t", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    return path


def export_scene2_csv(xlsx_path: Path, csv_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb[wb.sheetnames[0]]
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(list(row))


def write_audit(repo_root: Path, changes: Sequence[Mapping[str, Any]]) -> None:
    output_dir = repo_root / "analysis_outputs" / "gt_completion"
    output_dir.mkdir(parents=True, exist_ok=True)
    mapping_rows = []
    for scene, rows in CONFIRMED_CANONICAL.items():
        for row_index, canonical in rows.items():
            mapping_rows.append({
                "unified_id": f"{scene}::{row_index}", "scene": scene,
                "row_index": row_index, "canonical_referents": ", ".join(canonical),
            })
    mapping_rows.sort(key=lambda row: (row["scene"], int(row["row_index"])))
    with (output_dir / "confirmed_gt_mapping.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(mapping_rows[0]), lineterminator="\n")
        writer.writeheader()
        writer.writerows(mapping_rows)
    with (output_dir / "data_repair_changes.csv").open("w", encoding="utf-8", newline="") as handle:
        fields = ["file", "sheet", "excel_row", "column", "old_value", "new_value"]
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows(changes)
    alias_rows = [
        {"scene": scene, "canonical_anchor": canonical, "alias": alias}
        for scene, mappings in ALIASES.items()
        for canonical, aliases in mappings.items()
        for alias in aliases
    ]
    with (output_dir / "alias_additions.csv").open("w", encoding="utf-8", newline="") as handle:
        fields = ["scene", "canonical_anchor", "alias"]
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows(alias_rows)
    payload = {
        "confirmed_mapping_count": len(mapping_rows),
        "scene2_extra_exp3_repair": {"unified_id": "scene2::569", "canonical_referents": ["drawer1"]},
        "new_canonical_anchors": [{"scene": "scene2", "object_name": "drawer2", "point": [0.638, -1.227, 5.241]}],
        "forbidden_new_anchors": ["laptop3", "Printer1", "heater", "ceiling", "room1_front_area", "centerpiece"],
        "changes_this_invocation": len(changes),
        "alias_addition_count": len(alias_rows),
    }
    (output_dir / "data_repair_summary.json").write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--repo_root", default=".")
    args = parser.parse_args()
    repo_root = Path(args.repo_root).resolve()
    changes: List[Dict[str, Any]] = []

    workbook_paths = {
        "scene2": "scene2_cleaned_v2.xlsx",
        "scene4_room1": "scene4_room1.xlsx",
        "scene4_room3": "scene4_room3.xlsx",
        "scene4_room4": "scene4_room4.xlsx",
        "scene5": "scene5.xlsx",
    }
    for scene, filename in workbook_paths.items():
        repair_workbook(repo_root, scene, filename, changes)
    for scene in ALIASES:
        repair_anchor_table(repo_root, scene, changes)
    export_scene2_csv(repo_root / "data/scene2_cleaned_v2.xlsx", repo_root / "data/scene2_cleaned_v2.csv")
    write_audit(repo_root, changes)
    print(json.dumps({"confirmed_gt": 28, "changes": len(changes), "new_anchor": "scene2:drawer2"}))


if __name__ == "__main__":
    main()
